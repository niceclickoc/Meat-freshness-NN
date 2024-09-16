import pandas as pd
import os

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def generate_report(supplier_number, total_meat, fresh_meat, half_fresh_meat, spoiled_meat, spoiled_meat_images, output_file):

    current_date = datetime.now().strftime("%Y-%m-%d")

    report_data = {
        "Номер поставщика": [supplier_number],
        "Дата": [current_date],
        "Всего мяса": [total_meat],
        "Свежего мяса": [fresh_meat],
        "Полу-свежего мяса": [half_fresh_meat],
        "Испорченного мяса": [spoiled_meat],
        "Ссылки на испорченное мясо": ['\n'.join(spoiled_meat_images)]
    }


    df = pd.DataFrame(report_data)

    if not os.path.exists(output_file):

        df.to_excel(output_file, index=False, sheet_name="Отчет")

        workbook = load_workbook(output_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=7, max_col=7):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Ширина столбцов
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = max_length + 2  # +2 для запаса

        # Высота строк
        for row in sheet.iter_rows(min_row=2):
            sheet.row_dimensions[row[0].row].height = 15

        workbook.save(output_file)
        print("="*121)
        print(f"Создан новый отчет: {output_file}")

    else:

        workbook = load_workbook(output_file)
        sheet = workbook["Отчет"]

        row_index = sheet.max_row + 1
        for r in dataframe_to_rows(df, index=False, header=False):
            sheet.append(r)

        for row in sheet.iter_rows(min_row=row_index, max_row=row_index, min_col=7, max_col=7):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Ширина столбцов
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = max_length + 2  # +2 для запаса

        # Высота строк
        for row in sheet.iter_rows(min_row=row_index):
            sheet.row_dimensions[row[0].row].height = 15

        workbook.save(output_file)
        print("="*121)
        print(f"Данные добавлены в существующий отчет: {output_file}")
