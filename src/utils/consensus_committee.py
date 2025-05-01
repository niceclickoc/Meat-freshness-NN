import numpy as np
from openpyxl import load_workbook


# Import data from Excel report
def load_data_from_excel(file):
    """
    Loading quantity of fresh, half-fresh and spoiled meat to calculate the supplier's coefficient
    For now we take only the last stroke data

    :return: quantity of fresh, half-fresh and spoiled meat
    """
    try:
        workbook = load_workbook(file)
        sheet = workbook.active

        def get_last_row_with_data(sheet):
            for row in range(sheet.max_row, 0, -1):
                if any(cell.value is not None for cell in sheet[row]):
                    return row
            return None

        last_row = get_last_row_with_data(sheet)

        fresh = sheet.cell(row=last_row, column=4).value
        half_fresh = sheet.cell(row=last_row, column=5).value
        spoiled = sheet.cell(row=last_row, column=6).value
        return fresh, half_fresh, spoiled

    except FileNotFoundError:
        print("Файл не найден или ещё не создан")
        return None, None, None

    except Exception as e:
        print("Ошибка", e)
        return None, None, None


class ConsensusCommittee:
    def __init__(self, weights, agent_coeffs, threshold_auto_confirm=0.95, threshold_consensus_lower=0.62,
                 threshold_consensus_upper=0.8, threshold_pred_model=0.85):
        """
        Initialize the committee with weights and agent coefficients.

        :param weights: list of weights for chromatic, HOG, and depth models [w1, w2, w3]
        :param agent_coeffs: list of significance coefficients for each agent [A1, A2, A3]
        :param threshold_auto_confirm: threshold for automatic defect confirmation
        :param threshold_consensus_lower: lower bound threshold for consensus process
        :param threshold_consensus_upper: upper bound threshold for consensus process
        :param threshold_pred_model: threshold for probabilistic prediction model
        """
        self.weights = weights
        self.agent_coeffs = agent_coeffs
        self.threshold_auto_confirm = threshold_auto_confirm
        self.threshold_consensus_lower = threshold_consensus_lower
        self.threshold_consensus_upper = threshold_consensus_upper
        self.threshold_pred_model = threshold_pred_model

    def sigmoid(self, x):
        """Sigmoid function for probability scaling."""
        return 1 / (1 + np.exp(-x))

    def weighted_average(self, chromatic_prob, hog_prob, depth_prob):
        """
        Calculates the weighted average of defect probabilities using the provided formula:

        Pvoc = Q(Pdef) = (ω1 * A1 * Pdef1 + ω2 * A2 * Pdef2 + ω3 * A3 * Pdef3) / (ω1 + ω2 + ω3)

        :param chromatic_prob: probability from chromatic model
        :param hog_prob: probability from HOG model
        :param depth_prob: probability from depth map model
        :return: weighted average of probabilities
        """
        numerator = (self.weights[0] * self.agent_coeffs[0] * chromatic_prob +
                     self.weights[1] * self.agent_coeffs[1] * hog_prob +
                     self.weights[2] * self.agent_coeffs[2] * depth_prob)
        denominator = sum(self.weights)
        weighted_avg = numerator / denominator

        if isinstance(weighted_avg, np.ndarray):
            weighted_avg = np.mean(weighted_avg)
        return weighted_avg

    def evaluate(self, chromatic_prob, hog_prob, depth_prob, pred_prob=None):
        """
        Evaluate defect probability based on agent results using all provided thresholds.

        :param chromatic_prob: probability from chromatic model
        :param hog_prob: probability from HOG model
        :param depth_prob: probability from depth model
        :param pred_prob: probability from probabilistic prediction model (optional)
        :return: evaluation result and final probability
        """
        # Step 1: Calculate the weighted average
        weighted_avg = self.weighted_average(chromatic_prob, hog_prob, depth_prob)
        result = self.sigmoid(weighted_avg)

        # Step 2: Check automatic defect confirmation
        if result >= self.threshold_auto_confirm:
            return "Defect Confirmed", result

        # Step 3: Human needed due high k
        elif self.threshold_consensus_upper < result < self.threshold_auto_confirm:
            return "Human needed", result

        # Step 4: Check if consensus process is needed
        elif self.threshold_consensus_lower <= result <= self.threshold_consensus_upper:
            print("="*40)
            print("Вероятность испорченности", result)
            fresh, half_fresh, spoiled = load_data_from_excel('./results/report.xlsx')

            def supplier_coef(fresh, half_fresh, spoiled):
                if (fresh and half_fresh and spoiled) is not None:
                    a = (fresh * 1) + (half_fresh * 0.5) + (spoiled * 0)
                    b = a / sum([fresh, half_fresh, spoiled])
                    print('Коэффициент поставщика:',b)
                    return b
                else:
                    return 0.9

            try:
                final_result = result * (1 / supplier_coef(fresh, half_fresh, spoiled))
            except ZeroDivisionError:
                final_result = result

            print('Вероятность с коэффициентом', final_result)
            print("=" * 40)
            if final_result <= self.threshold_consensus_lower:
                return "Ok", final_result
            if final_result <= self.threshold_consensus_upper:
                return "Human needed", final_result
            else:
                return "Defect Confirmed", final_result

        # Step 5: Check if probabilistic prediction model should be used
        # elif result <= self.threshold_pred_model:
        #     if pred_prob is not None:
        #         # Apply formula for probabilistic prediction model integration:
        #         final_result = self.sigmoid(result + self.weights[3] * pred_prob)
        #         if final_result <= self.threshold_auto_confirm:
        #             return "Human Expert Needed", final_result
        #         else:
        #             return "Defect Confirmed with Prediction", final_result
        #     else:
        #         return "Human Expert Needed", result

        # Step 6: No defect detected
        return "No Defect", result
